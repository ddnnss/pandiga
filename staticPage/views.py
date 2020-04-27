import json
from .models import City
from django.http import HttpResponseRedirect, JsonResponse
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import load_workbook

def type(request):
    wb = load_workbook(filename='C:/Users/ххх/PycharmProjects/ideaHome/cats.xlsx')
    sheet = wb.active

    max_row = sheet.max_row

    max_column = sheet.max_column
    for i in range(1, max_row + 1):
        # worksheet.write('A{}'.format(row), cat_id)
        # worksheet.write('B{}'.format(row), cat_name)
        # worksheet.write('C{}'.format(row), cat_parent_id)
        # worksheet.write('D{}'.format(row), cat_description)
        # worksheet.write('E{}'.format(row), cat_img)
        # worksheet.write('F{}'.format(row), cat_title)
        # worksheet.write('G{}'.format(row), cat_keywords)
        # worksheet.write('H{}'.format(row), cat_meta_description)
        old_id = sheet.cell(row=i, column=1).value
        cat_name = sheet.cell(row=i, column=2).value
        cat_parent_id = sheet.cell(row=i, column=3).value
        try:
            cat_description = sheet.cell(row=i, column=4).value.replace('_x000D_', '')
        except:
            cat_description = sheet.cell(row=i, column=4).value
        cat_img = sheet.cell(row=i, column=5).value
        cat_title = sheet.cell(row=i, column=6).value
        cat_keywords = sheet.cell(row=i, column=7).value
        cat_meta_description = sheet.cell(row=i, column=8).value
        # if cat_parent_id ==0:
        #     Category.objects.create(old_id=old_id,name=cat_name,description=cat_description,image='images/catalog/categories/'+cat_img,
        #                             page_title=cat_title,page_description=cat_meta_description,page_keywords=cat_keywords)
        if cat_parent_id != 0:
            cat = Category.objects.get(old_id=cat_parent_id)
            SubCategory.objects.create(old_id=old_id, name=cat_name, description=cat_description, category=cat,
                                       page_title=cat_title, page_description=cat_meta_description,
                                       page_keywords=cat_keywords)
    return render(request, 'page/about.html', locals())


def create_city(request):
    from .sity_search import cities

    for i in cities:
        City.objects.create(city=i['city'],region=i['region'])
        print(i)

def get_city(request):
    request_unicode = request.body.decode('utf-8')
    request_body = json.loads(request_unicode)
    print(request_body)
    cities = City.objects.filter(city__startswith=request_body['query'].capitalize())

    return_dict = list()
    for i in cities:
        return_dict.append({'id' : i.id, 'city': i.city, 'region': i.region})
    return JsonResponse(return_dict, safe=False)


def index(request):
    indexPage = True
    return render(request, 'staticPage/index.html', locals())


def login_page(request):
    if not request.user.is_authenticated:
        loginActive = 'menu-link-active '
        return render(request, 'staticPage/login.html', locals())
    else:
        return HttpResponseRedirect('/')

def robots(request):
    robotsTxt = f"User-agent: *\nDisallow: /admin/\nHost: https://www.pandiga.ru/\nSitemap: https://www.pandiga.ru/sitemap.xml"
    return HttpResponse(robotsTxt, content_type="text/plain")

def about(request):
    return render(request, 'staticPage/about.html', locals())
def forpartners(request):
    return render(request, 'staticPage/forpartners.html', locals())
def servicerules(request):
    return render(request, 'staticPage/servicerules.html', locals())
def fortechniqueowners(request):
    return render(request, 'staticPage/forperformers.html', locals())
def tariff(request):
    return render(request, 'staticPage/tariff.html', locals())
def questions(request):
    return render(request, 'staticPage/questions.html', locals())
def contacts(request):
    return render(request, 'staticPage/contacts.html', locals())
def licenzionnoe_soglashenie(request):
    return render(request, 'staticPage/servicerules-licenzionnoe_soglashenie.html', locals())

def privacypolicy(request):
    return render(request, 'staticPage/servicerules-privacypolicy.html', locals())

def publichnyj_agentskij_dogovor(request):
    return render(request, 'staticPage/servicerules-publichnyj_agentskij_dogovor.html', locals())
def usloviya_ispolzovaniya(request):
    return render(request, 'staticPage/servicerules-usloviya_ispolzovaniya.html', locals())
def vacancies(request):
    return render(request, 'staticPage/vacancies.html', locals())